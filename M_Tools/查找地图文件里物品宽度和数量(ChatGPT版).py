import os
import json
from openpyxl import Workbook

# 指定文件夹路径
folder_path = r'D:\set_excel\08level'

# 创建一个 Excel 文件
wb = Workbook()
ws = wb.active
ws['A1'] = 'File Name'
ws['B1'] = 'Width'
ws['C1'] = 'Num'

# 存储每个文件的数据
file_data = {}

# 遍历文件夹中的所有文件
for filename in os.listdir(folder_path):
    if filename.endswith('.json'):
        file_path = os.path.join(folder_path, filename)
        with open(file_path, 'r') as f:
            data = json.load(f)
            file_name = filename[:-5]  # 去除文件扩展名
            for item in data.get('numList', []):
                width = item.get('width')
                num = item.get('num')
                if width is not None:
                    if file_name not in file_data:
                        file_data[file_name] = {'width': [], 'num': []}
                    file_data[file_name]['width'].append(width)
                if num is not None:
                    if file_name not in file_data:
                        file_data[file_name] = {'width': [], 'num': []}
                    file_data[file_name]['num'].append(num)

# 将数据写入Excel文件
row_num = 2
for file_name, data in file_data.items():
    widths = data['width']
    nums = data['num']
    for width, num in zip(widths, nums):
        ws[f'A{row_num}'] = file_name
        ws[f'B{row_num}'] = width
        ws[f'C{row_num}'] = num
        row_num += 1

# 保存 Excel 文件
excel_filename = 'output.xlsx'
wb.save(excel_filename)
print(f'Excel file "{excel_filename}" has been generated.')
