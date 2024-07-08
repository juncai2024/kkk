import random
import openpyxl

data = [
    ["C0100", "C0101", "C0102", "C0103", "C0104", "C0105", "C0106"],
    ["C0107", "C0108", "C0109", "C0110", "C0111", "C0112", "C0113", "C0114", "C0115", "C0116", "C0117", "C0118", "C0119", "C0120", "C0121", "C0122", "C0123", "C0124", "C0125", "C0126", "C0127", "C0128", "C0129", "C0130", "C0131"],
    ["C0132", "C0133", "C0134", "C0135", "C0136", "C0137", "C0138", "C0139", "C0140", "C0141", "C0142", "C0143", "C0144", "C0145", "C0146", "C0147"],
    ["C0148", "C0149", "C0150", "C0151", "C0152", "C0153", "C0154", "C0155", "C0156", "C0157", "C0158", "C0159", "C0160", "C0161", "C0162", "C0163"],
    ["C0164", "C0165", "C0166", "C0167", "C0168", "C0169", "C0170", "C0171"]
]

weights = [230, 320, 215, 115, 25]

def weighted_choice(choices):
    total = sum(weight for _, weight in choices)
    r = random.uniform(0, total)
    upto = 0
    for choice, weight in choices:
        if upto + weight >= r:
            return choice
        upto += weight

def random_element(arr):
    return random.choice(arr)

# 计算每个数组的抽取概率
total_weight = sum(weights)
probabilities = [(i, weight / total_weight) for i, weight in enumerate(weights)]

# 输入重复次数和每次运行抽取次数
total_runs = int(input("请输入运行次数: "))
runs_per_iteration = int(input("请输入每次运行抽取次数: "))

# 创建一个新的 Excel 工作簿
wb = openpyxl.Workbook()

# 进行总体的循环
for run_num in range(total_runs):
    print(f"\n第{run_num + 1}次运行:")

    # 创建一个新的工作表，命名为 Run_1, Run_2, ...
    ws = wb.create_sheet(title=f"Run_{run_num + 1}")

    # 进行每次运行的循环
    for col_num in range(runs_per_iteration):
        # 根据概率抽取数组
        selected_array_index = weighted_choice(probabilities)
        selected_array = data[selected_array_index]

        # 从选定的数组中完全随机抽取一个值
        selected_value = random_element(selected_array)

        # 将结果输出到控制台
        print(selected_value)

        # 将结果写入 Excel 表的不同列
        ws.cell(row=col_num + 1, column=1, value=selected_value)

# 删除默认创建的第一个工作表
del wb['Sheet']

# 保存 Excel 文件
wb.save("抽取结果.xlsx")
