import os
import json
from openpyxl import Workbook

# 指定文件夹路径
folder_path = r'D:\set_excel\08level'

# 创建一个 Excel 文件
wb = Workbook()
ws = wb.active
ws['A1'] = 'File Name'
ws['B1'] = 'Badge Value'
row_num = 2

# 遍历文件夹中的所有文件
for filename in os.listdir(folder_path):
    if filename.endswith('.json'):
        file_path = os.path.join(folder_path, filename)
        with open(file_path, 'r') as f:
            data = json.load(f)
            if 'badge' in data and data['badge'] in ['hard', 'extreme']:
                ws[f'A{row_num}'] = filename
                ws[f'B{row_num}'] = data['badge']
                row_num += 1

# 保存 Excel 文件
excel_filename = 'badge_values.xlsx'
wb.save(excel_filename)
print(f'Excel file "{excel_filename}" has been generated.')
